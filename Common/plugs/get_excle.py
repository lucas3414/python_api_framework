import os, sys
from openpyxl import load_workbook
from Common.plugs.get_config import r_config
from Common.plugs.get_log import Log

BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(__file__)))

if sys.platform == "win32":
    ENV_CONF_DIR = os.path.join(BASE_DIR, 'Common/conf/env_config.ini').replace('/', '\\')
else:
    ENV_CONF_DIR = os.path.join(BASE_DIR, 'Common/conf/env_config.ini')

log_path = r_config(ENV_CONF_DIR, "log", "log_path")
test_case_path = r_config(ENV_CONF_DIR, 'test_case', 'test_case_path')

logger = Log(log_path)


class DoExcle:

    def __init__(self, filename, sheetname):
        self.filename = filename
        self.sheetname = sheetname

    def read_data(self):
        test_data = []
        try:
            logger.info("准备开始打开 {0} 文件".format(self.filename))
            wb = load_workbook(self.filename)
        except Exception as e:
            logger.error("打开 {0} 文件失败: {1}".format(self.filename, str(e)))
        try:
            logger.info("准备开始打开 {0}-{1} sheet页".format(self.filename, self.sheetname))
            sheet = wb[self.sheetname]
        except Exception as e:
            logger.error("读取 {0}-{1} sheet页数据失败: {2}".format(self.filename, self.sheetname, str(e)))

        logger.info("开始读取数据")
        for i in range(2, sheet.max_row + 1):
            row_data = {}
            row_data['sheetname'] = self.sheetname
            row_data['id'] = sheet.cell(i, 1).value
            row_data['level'] = sheet.cell(i, 2).value
            row_data['description'] = sheet.cell(i, 3).value
            row_data['method'] = sheet.cell(i, 4).value
            row_data['url'] = sheet.cell(i, 5).value
            row_data['parm'] = sheet.cell(i, 6).value
            row_data['excepted'] = sheet.cell(i, 7).value
            row_data['sql'] = sheet.cell(i, 8).value
            test_data.append(row_data)
        logger.info("读取 {0}-{1}-{2}-{3}-{4}-{5}-{6}-{7}-{8}-{9} sheet页数据成功".format(
            self.filename, self.sheetname, 'id', 'level', 'description', 'method', 'url', 'parm', 'excepted', 'sql'
        ))
        logger.info("读取数据结束")
        return test_data

    def write_data(self, sheetname, row, col, value):
        try:
            logger.info("准备写入数据".format(self.filename))
            logger.info("准备开始打开 {0} 文件".format(self.filename))
            wb = load_workbook(self.filename)
        except Exception as e:
            logger.error("打开 {0} 文件失败: {1}".format(self.filename, str(e)))
        try:
            logger.info("准备开始打开 {0}-{1} sheet页".format(self.filename, self.sheetname))
            sheet = wb[self.sheetname]
        except Exception as e:
            logger.error("读取 {0}-{1} sheet页数据失败: {2}".format(self.filename, self.sheetname, str(e)))

        logger.info("准备写入数据")
        logger.info("sheet:{0}, 行数：{1}， 列数:{2}, 值:{3}".format(sheetname, row, col, value))
        try:
            sheet.cell(row, col).value = value
            wb.save(self.filename)
            logger.info("写入数据成功")
        except Exception as e:
            logger.info("写入数据失败:{0}".format(str(e)))


def split_list(lst, n=3):
    data_split_list = [[] for x in range(n)]
    [data_split_list[0].append(i) for i in lst if i['level'] == 'P1']
    [data_split_list[1].append(i) for i in lst if i['level'] == 'P2']
    [data_split_list[2].append(i) for i in lst if i['level'] == 'P3']
    return data_split_list


if __name__ == '__main__':
    DE = DoExcle(test_case_path, 'user')
    caseList = split_list(DE.read_data())
    # print(caseList[0])
    # print(caseList[0][0]['sql'])
    pass
